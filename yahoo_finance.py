### import packages

import yfinance as yf
from openpyxl import load_workbook

### loading workbook

wb = load_workbook('yahoo.xlsx')
ws = wb.active

### defining maximum column and row

def get_maximum_cols():
    for i in range(2, 20000):
        if ws.cell(row=1, column= i).value == None:
            max_col = i
            break
    return max_col-1

def get_maximum_rows():
    for i in range(2, 20000):
        if ws.cell(row=i, column= 1).value == None:
            max_row = i
            break
    return max_row-1

### get ticker list from the first row

ticker_list = []

for row in ws.iter_rows(min_row=1, max_row=1, min_col=2, max_col=get_maximum_cols()):
    for cell in row:
        ticker_list.append(cell.value)

#print(ticker_list)

### load tickers

loaded_tickers = []

for ticker in ticker_list:
    obj = yf.Ticker(ticker)
    loaded_tickers.append(obj)

### get data for each ticker

all_ticker_data = []

for i in range(0, len(loaded_tickers)):
    all_ticker_data.extend((loaded_tickers[i].info.get('currentPrice', 'N/A'), loaded_tickers[i].info.get('trailingPE', 'N/A'), loaded_tickers[i].info.get('forwardPE', 'N/A'), loaded_tickers[i].info.get('trailingEps', 'N/A'),
                        loaded_tickers[i].info.get('beta', 'N/A'), loaded_tickers[i].info.get('fiftyTwoWeekHigh', 'N/A'), loaded_tickers[i].info.get('fiftyTwoWeekLow', 'N/A'), loaded_tickers[i].info.get('fiftyDayAverage', 'N/A'),
                        loaded_tickers[i].info.get('twoHundredDayAverage', 'N/A'), loaded_tickers[i].info.get('dividendRate', 'N/A'), loaded_tickers[i].info.get('dividendYield', 'N/A')))

### put data into worksheet

i = 0

for col in ws.iter_cols(min_row=2, max_row=get_maximum_rows(), min_col=2, max_col=get_maximum_cols()):
    for cell in col:
        cell.value = all_ticker_data[i]
        i += 1

### save workbook

wb.save('yahoo.xlsx')


